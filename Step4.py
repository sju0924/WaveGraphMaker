import pandas as pd
import numpy as np
from math import sqrt
from Step2 import plot_response, plot_scaleup_factor, plot_acceleration,get_scaleup_factor
from openpyxl import Workbook
PATH_STEP4 = "./Step4"
delta_T = 0.02
def get_srss_table():
    
    # Create a DataFrame
    df = pd.DataFrame([], columns=['T'])

    # 응답스펙트럼 X,Y 방향 테이블 불러오기
    path = PATH_STEP4 + "/input/SGS_result"
    for i in range(100):
        idx = str(i+1)
        eq_filename_x = path + "/"+idx+"x.sgs"
        eq_filename_y = path + "/"+idx+"y.sgs"       
        try:
            with open(eq_filename_x,'r',encoding='utf-8') as f:

                lines = f.read().splitlines()[9:-1]

                # x방향 데이터 불러오기기
                numeric_data = [line.split(',') for line in lines]
                df_x = pd.DataFrame(numeric_data, columns=['T', idx + 'x'])

                if len(numeric_data) > len(df['T']):
                    df['T'] = pd.to_numeric(df_x['T'])

                df[idx + 'x'] = pd.to_numeric(df_x[idx + 'x'])


            with open(eq_filename_y,'r',encoding='utf-8') as f:
                
                lines = f.read().splitlines()[9:-1]

                # x방향 데이터 불러오기기
                numeric_data = [line.split(',') for line in lines]
                df_y = pd.DataFrame(numeric_data, columns=['T', idx + 'y'])

                if len(numeric_data) > len(df['T']):
                    df['T'] = pd.to_numeric(df_y['T'])  

                df[idx + 'y'] = pd.to_numeric(df_y[idx + 'y'])
            
            # srss 계산산
            df[str(i+1)+'srss'] = df[str(i+1)+'x'] * df[str(i+1)+'x'] + df[str(i+1)+'y'] * df[str(i+1)+'y']
            df[str(i+1)+'srss'] = df[str(i+1)+'srss'].apply(sqrt)      

        except FileNotFoundError:
            break
        except Exception as e:
            print(e)

    df.set_index('T', inplace=True)

    return df

# @input    srss_table: x,y, srss 컬럼을 포함한 데이터프레임
# @output   Step2/SRSS 위치에 srss table의 그래프 이미지 파일 저장
def plot_srss(srss_table):
    plot_response(srss_table, PATH_STEP4 + "/output/SRSS")


def get_acceleration_table():
    # Create a DataFrame
    df = pd.DataFrame()

    # 응답스펙트럼 X,Y 방향 테이블 불러오기
    path = PATH_STEP4 + "/input/RSP_match_result"
    for i in range(100):
        idx = str(i+1)    
        acc_filename_x = path + "/"+idx+"x.txt"
        acc_filename_y = path + "/"+idx+"y.txt"       
        cur='x'
        try:
            with open(acc_filename_x,'r',encoding='utf-8') as f:
                
                lines = f.read().splitlines()

                # x방향 데이터 불러오기
                raw_data = [line.split() for line in lines]
                dt = float(raw_data[1][1])
                numeric_data = []

                # 세 번째 줄부터 모든 숫자를 1차원 리스트로 추출
                for row in raw_data[2:]:
                    numeric_data.extend([float(x.replace('E', 'e')) for x in row]) 
                
                # 시간 배열 생성
                time_array = np.arange(0, dt * len(numeric_data), dt)
                time_array = time_array[:len(numeric_data)]

                # DataFrame 생성
                df_x = pd.DataFrame({
                    'T': time_array,
                     idx + 'x': numeric_data
                })
                df_x = df_x[np.isclose(df_x['T'] % delta_T, 0, atol=1e-10)]
            
                df[idx + 'xT'] = pd.to_numeric(df_x['T'])
                df[idx + 'xAcc'] = pd.to_numeric(df_x[idx + 'x'])

            with open(acc_filename_y,'r',encoding='utf-8') as f:
                cur='y'
                lines = f.read().splitlines()

                # y방향 데이터 불러오기
                raw_data = [line.split() for line in lines]
                dt = float(raw_data[1][1])
                numeric_data = []

                # 세 번째 줄부터 모든 숫자를 1차원 리스트로 추출
                for row in raw_data[2:]:
                    numeric_data.extend([float(x.replace('E', 'e')) for x in row]) 

                # 시간 배열 생성
                time_array = np.arange(0, dt * len(numeric_data), dt)
                time_array = time_array[:len(numeric_data)]

                # DataFrame 생성
                df_y = pd.DataFrame({
                    'T': time_array,
                     idx + 'y': numeric_data
                })
                df_y = df_y[np.isclose(df_y['T'] % delta_T, 0, atol=1e-10)]
            
                df[idx + 'yT'] = pd.to_numeric(df_y['T'])
                df[idx + 'yAcc'] = pd.to_numeric(df_y[idx + 'y'])
        except FileNotFoundError:
            break

        except Exception as e:
            print("idx: ",idx,", ",cur)
            print(e.__class__.__name__ , e)
    df = df.reset_index(drop=True)
    return df

def Step4_SF():
    srss_table = get_srss_table()
    scale = get_scaleup_factor(srss_table, new = 1)
    return srss_table, scale

def Step4_srss(srss_table, scale):
    plot_srss(srss_table)
    plot_scaleup_factor(srss_table, scale, PATH_STEP4 + "/output/SRSS_Scale", mod='total')

def Step4_acc(scale):
    acc_table = get_acceleration_table()
    plot_acceleration(acc_table, scale, PATH_STEP4 + "/output/Acceleration/2400", mod='text')

    acc_table_copy = get_acceleration_table()
    for i in range(100):
        idx = str(i+1)
        try:
            acc_table_copy[idx+"xT"] = acc_table_copy[idx+"xT"]
            acc_table_copy[idx+"yT"] = acc_table_copy[idx+"yT"]
            acc_table_copy[idx+"xAcc"] = acc_table_copy[idx+"xAcc"] * 4 / 5
            acc_table_copy[idx+"yAcc"] = acc_table_copy[idx+"yAcc"] * 4 / 5
        except KeyError:
            break
    plot_acceleration(acc_table_copy, scale, PATH_STEP4 + "/output/Acceleration/1000", mod = 'text')

    for i in range(100):
        idx = str(i+1)
        try:
            acc_table[idx+"xAcc(1400)"] = acc_table[idx+"xAcc"] * 4 / 5
            acc_table[idx+"yAcc(1400)"] = acc_table[idx+"yAcc"] * 4 / 5
            acc_table[idx+"xT(1400)"] = acc_table[idx+"xT"]
            acc_table[idx+"yT(1400)"] = acc_table[idx+"yT"]
        except KeyError:
            break

    acc_table = acc_table[[f"{i}{suffix}" for i in range(1, int(idx)) for suffix in ["xT","xAcc","xT(1400)","xAcc(1400)", "yT","yAcc","yT(1400)", "yAcc(1400)"]]]
    file_name = PATH_STEP4 + "/output/Acceleration/output.xlsx"
    workbook = Workbook()  # 파일이 없으면 새로 생성

    # 시트 선택 또는 새로 생성
    sheet_name = "Sheet1"
    worksheet = workbook.create_sheet(sheet_name)

    # 1️⃣ 첫 번째 행 (병합하여 "1x", "2x", ... "7x" 작성)
    for i in range(int(idx)-1):
        col_start = (i * 8) + 1  # openpyxl은 1-based index
        col_end = col_start + 3
        worksheet.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
        worksheet.cell(row=1, column=col_start, value=f"{i+1}x")

        col_start = (i * 8) + 5
        col_end = col_start + 3
        worksheet.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
        worksheet.cell(row=1, column=col_start, value=f"{i+1}y")

    # 2️⃣ 두 번째 행 (각 열에 2400, 1000 입력)
    for i in range((int(idx)-1) * 2):
        col_start = (i * 4)+1
        worksheet.cell(row=2, column=col_start, value=2400)
        worksheet.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_start+1)
        worksheet.cell(row=2, column=col_start + 2, value=1000)
        worksheet.merge_cells(start_row=2, start_column=col_start+2, end_row=2, end_column=col_start+3)

    # 엑셀 파일 저장
    workbook.save(file_name)
    workbook.close()

    with pd.ExcelWriter(file_name, engine="openpyxl",mode='a',  if_sheet_exists="overlay") as writer:
        acc_table.to_excel(writer, sheet_name="Sheet1", startrow=2, index=False,header=False)  # 3번째 행부터 입력

def main():
    cmd = ""    
    while(cmd != "yes"):
        print("RSP match 결과값을 input 폴더에 넣어주세요")
        cmd = input("완료 후 'yes' 를 입력하세요 > ")

    acc_table = get_acceleration_table()

    cmd = ""    
    while(cmd != "yes"):
        print("SGS 결과값을 input 폴더에 넣어주세요")
        cmd = input("완료 후 'yes' 를 입력하세요 > ")

    srss_table = get_srss_table()
    scale = get_scaleup_factor(srss_table, new = 1)

    plot_srss(srss_table)
    plot_scaleup_factor(srss_table, scale, PATH_STEP4 + "/output/SRSS_Scale")
    plot_acceleration(acc_table, scale, PATH_STEP4 + "/output/Acceleration/2400")

    acc_table_copy = acc_table
    for i in range(100):
        idx = str(i+1)
        try:
            acc_table_copy[idx+"xAcc"] = acc_table[idx+"xAcc"] * 0.8
            acc_table_copy[idx+"yAcc"] = acc_table[idx+"yAcc"] * 0.8
        except KeyError:
            break
    plot_acceleration(acc_table, scale, PATH_STEP4 + "/output/Acceleration/1000")

    for i in range(100):
        idx = str(i+1)
        try:
            acc_table[idx+"xAcc(1400)"] = acc_table[idx+"xAcc"] * 0.8
            acc_table[idx+"yAcc(1400)"] = acc_table[idx+"yAcc"] * 0.8
            acc_table[idx+"xT(1400)"] = acc_table[idx+"xT"]
            acc_table[idx+"yT(1400)"] = acc_table[idx+"yT"]
        except KeyError:
            break

    acc_table = acc_table[[f"{i}{suffix}" for i in range(1, int(idx)) for suffix in ["xT","xAcc","xT(1400)","xAcc(1400)", "yT","yAcc","yT(1400)", "yAcc(1400)"]]]
    file_name = PATH_STEP4 + "/output/Acceleration/output.xlsx"
    workbook = Workbook()  # 파일이 없으면 새로 생성

    # 시트 선택 또는 새로 생성
    sheet_name = "Sheet"
    worksheet = workbook.create_sheet(sheet_name)

    # 1️⃣ 첫 번째 행 (병합하여 "1x", "2x", ... "7x" 작성)
    for i in range(int(idx)-1):
        col_start = (i * 8) + 1  # openpyxl은 1-based index
        col_end = col_start + 3
        worksheet.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
        worksheet.cell(row=1, column=col_start, value=f"{i+1}x")

        col_start = (i * 8) + 5
        col_end = col_start + 3
        worksheet.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_end)
        worksheet.cell(row=1, column=col_start, value=f"{i+1}y")

    # 2️⃣ 두 번째 행 (각 열에 2400, 1000 입력)
    for i in range((int(idx)-1) * 2):
        col_start = (i * 4)+1
        worksheet.cell(row=2, column=col_start, value=2400)
        worksheet.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_start+1)
        worksheet.cell(row=2, column=col_start + 2, value=1000)
        worksheet.merge_cells(start_row=2, start_column=col_start+2, end_row=2, end_column=col_start+3)

    # 엑셀 파일 저장
    workbook.save(file_name)
    workbook.close()

    with pd.ExcelWriter(file_name, engine="openpyxl",mode='a',  if_sheet_exists="overlay") as writer:
        acc_table.to_excel(writer, sheet_name="Sheet", startrow=2, index=False)#,header=False)  # 3번째 행부터 입력

    # ToDo: 1400년 주기 지진파 데이터 작성하기

if __name__=="__main__":
    main()